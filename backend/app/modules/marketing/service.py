from __future__ import annotations
import csv
import io
from uuid import UUID

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.marketing.models import Lead, LeadStatus, LeadTag, Source
from app.modules.marketing.schemas import (
    LeadCreate,
    LeadImportResult,
    LeadStatisticsResponse,
    LeadStatisticsRow,
    LeadStatusResponse,
    LeadUpdate,
)
from app.shared.admin_crud import CRUDService, OrgScope
from app.shared.exceptions import ValidationError


class LeadService(CRUDService[Lead]):
    def __init__(self, db: AsyncSession):
        super().__init__(Lead, db)

    async def _load_tags(self, tag_ids: list[UUID]) -> list[LeadTag]:
        if not tag_ids:
            return []
        tags = (
            await self.db.execute(select(LeadTag).where(LeadTag.id.in_(tag_ids)))
        ).scalars().all()
        return list(tags)

    async def create_lead(self, data: LeadCreate, manager_id: UUID) -> Lead:
        payload = data.model_dump(exclude={"tag_ids"})
        payload.setdefault("user_id", None)
        if payload.get("user_id") is None:
            payload["user_id"] = manager_id
        lead = Lead(**payload)
        lead.tags = await self._load_tags(data.tag_ids)
        self.db.add(lead)
        await self.db.commit()
        await self.db.refresh(lead)
        return lead

    async def update_lead(self, lead_id: UUID, data: LeadUpdate) -> Lead:
        lead = await self.get(lead_id)
        payload = data.model_dump(exclude_unset=True, exclude={"tag_ids"})
        for key, value in payload.items():
            setattr(lead, key, value)
        if data.tag_ids is not None:
            lead.tags = await self._load_tags(data.tag_ids)
        await self.db.commit()
        await self.db.refresh(lead)
        return lead

    async def assign_tags(self, lead_id: UUID, tag_ids: list[UUID]) -> Lead:
        lead = await self.get(lead_id)
        lead.tags = await self._load_tags(tag_ids)
        await self.db.commit()
        await self.db.refresh(lead)
        return lead

    async def statistics(self, org_scope: OrgScope = None) -> LeadStatisticsResponse:
        """Воронка лидов: матрица «источник × статус» (ТЗ §6)."""
        statuses = (
            await self.db.execute(select(LeadStatus).order_by(LeadStatus.sort))
        ).scalars().all()

        query = (
            select(Lead.source_id, Lead.status_id, func.count())
            .where(Lead.deleted_at.is_(None))
            .group_by(Lead.source_id, Lead.status_id)
        )
        if org_scope is not None:
            query = query.where(Lead.organization_id.in_(org_scope))
        counts = (await self.db.execute(query)).all()

        sources = {
            s.id: s.name
            for s in (await self.db.execute(select(Source))).scalars().all()
        }

        matrix: dict[UUID | None, dict[str, int]] = {}
        for source_id, status_id, count in counts:
            row = matrix.setdefault(source_id, {})
            row[str(status_id)] = row.get(str(status_id), 0) + count

        rows = []
        totals: dict[str, int] = {}
        grand_total = 0
        for source_id, row_counts in matrix.items():
            row_total = sum(row_counts.values())
            grand_total += row_total
            for key, value in row_counts.items():
                totals[key] = totals.get(key, 0) + value
            rows.append(LeadStatisticsRow(
                source_id=source_id,
                source_name=sources.get(source_id, "Без источника"),
                counts=row_counts,
                total=row_total,
            ))
        rows.sort(key=lambda r: -r.total)

        return LeadStatisticsResponse(
            statuses=[LeadStatusResponse.model_validate(s) for s in statuses],
            rows=rows,
            totals=totals,
            grand_total=grand_total,
        )

    async def import_leads(self, filename: str, content: bytes, manager_id: UUID) -> LeadImportResult:
        """Импорт лидов из CSV/XLSX: customer_name, phone, type, quantity, comment."""
        rows: list[dict]
        if filename.lower().endswith(".csv"):
            text = content.decode("utf-8-sig")
            rows = list(csv.DictReader(io.StringIO(text)))
        elif filename.lower().endswith((".xlsx", ".xls")):
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise ValidationError("Импорт XLSX требует пакет openpyxl")
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            header = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            rows = [
                {header[i]: cell.value for i, cell in enumerate(row) if i < len(header)}
                for row in ws.iter_rows(min_row=2)
            ]
        else:
            raise ValidationError("Поддерживаются файлы .csv и .xlsx")

        created, errors = 0, []
        for index, row in enumerate(rows, start=2):
            name = str(row.get("customer_name") or row.get("name") or "").strip()
            if not name:
                errors.append(f"Строка {index}: пустое имя клиента")
                continue
            phone = str(row.get("phone") or "").strip()
            lead = Lead(
                customer_name=name,
                phones=[phone] if phone else [],
                type=str(row.get("type") or "offline"),
                quantity=int(row["quantity"]) if row.get("quantity") else None,
                comment=str(row.get("comment") or "") or None,
                user_id=manager_id,
            )
            self.db.add(lead)
            created += 1
        await self.db.commit()
        return LeadImportResult(created=created, errors=errors)
